# glpi_inventaire.py — Inventaire enrichi avec données réseau
# Projet CIMAT Béni Mellal

import requests
import socket
import subprocess
import platform
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
import os

# ── Config ────────────────────────────────────────────────
GLPI_URL   = "http://192.168.112.129/glpi"
APP_TOKEN  = "Qzqd9JxK0k9vcF9daXxT4zIJKZGht2r20pHgaYR4"
USER_TOKEN = "gMJqxaSdpy79c9HQout8reOo1PBY0UfM15FkQGYs"

# ── Connexion ─────────────────────────────────────────────
def connect_glpi():
    headers = {
        "Content-Type":  "application/json",
        "Authorization": f"user_token {USER_TOKEN}",
        "App-Token":     APP_TOKEN
    }
    r = requests.get(f"{GLPI_URL}/apirest.php/initSession", headers=headers)
    data = r.json()
    if isinstance(data, list) or "session_token" not in data:
        print(f" Erreur connexion : {data}")
        return None
    print(f" Connecté — Session : {data['session_token'][:12]}…")
    return data["session_token"]

def disconnect_glpi(session_token):
    headers = {
        "Content-Type":  "application/json",
        "Session-Token": session_token,
        "App-Token":     APP_TOKEN
    }
    requests.get(f"{GLPI_URL}/apirest.php/killSession", headers=headers)
    print(" Session fermée.")

# ── Récupération équipements GLPI ─────────────────────────
def get_computers(session_token):
    """Récupère tous les ordinateurs depuis GLPI."""
    headers = {
        "Content-Type":  "application/json",
        "Session-Token": session_token,
        "App-Token":     APP_TOKEN
    }
    params = {
        "expand_dropdowns": True,
        "range": "0-999",
        "with_networkports": True
    }
    r = requests.get(
        f"{GLPI_URL}/apirest.php/Computer",
        headers=headers,
        params=params
    )
    if r.status_code != 200:
        print(f" Erreur récupération : {r.text}")
        return []
    computers = r.json()
    if isinstance(computers, list) and len(computers) > 0:
        print(f"  {len(computers)} ordinateur(s) trouvé(s) dans GLPI.")
        return computers
    print("  Aucun ordinateur trouvé dans GLPI.")
    return []

def get_network_devices(session_token):
    """Récupère les équipements réseau (switches, routeurs)."""
    headers = {
        "Content-Type":  "application/json",
        "Session-Token": session_token,
        "App-Token":     APP_TOKEN
    }
    params = {"expand_dropdowns": True, "range": "0-999"}
    r = requests.get(
        f"{GLPI_URL}/apirest.php/NetworkEquipment",
        headers=headers,
        params=params
    )
    if r.status_code == 200:
        devices = r.json()
        if isinstance(devices, list):
            print(f" {len(devices)} équipement(s) réseau trouvé(s).")
            return devices
    return []

# ── Données réseau réelles ────────────────────────────────
def ping_host(ip):
    """Teste si un hôte répond au ping."""
    if not ip or ip == "N/A":
        return " Inconnu"
    try:
        param = "-n" if platform.system().lower() == "windows" else "-c"
        result = subprocess.run(
            ["ping", param, "1", "-w", "1000", ip],
            capture_output=True,
            text=True,
            timeout=3
        )
        if result.returncode == 0:
            return " En ligne"
        else:
            return " Hors ligne"
    except Exception:
        return " Inconnu"

def resolve_hostname(ip):
    """Résout le hostname depuis l'IP."""
    if not ip or ip == "N/A":
        return "N/A"
    try:
        hostname = socket.gethostbyaddr(ip)[0]
        return hostname
    except Exception:
        return "N/A"

def get_ip_from_glpi(computer, session_token):
    """Récupère l'IP d'un ordinateur via ses ports réseau."""
    headers = {
        "Content-Type":  "application/json",
        "Session-Token": session_token,
        "App-Token":     APP_TOKEN
    }
    computer_id = computer.get("id")
    try:
        r = requests.get(
            f"{GLPI_URL}/apirest.php/Computer/{computer_id}/NetworkPort",
            headers=headers
        )
        if r.status_code == 200:
            ports = r.json()
            if isinstance(ports, list):
                for port in ports:
                    ip = port.get("ip", "")
                    if ip and ip != "127.0.0.1":
                        return ip
    except Exception:
        pass
    return "N/A"

# ── Enrichissement des données ────────────────────────────
def enrichir_equipement(computer, session_token):
    """Ajoute les données réseau réelles à un équipement."""
    print(f"   Analyse : {computer.get('name', 'N/A')}…", end=" ")

    ip = get_ip_from_glpi(computer, session_token)

    # Si pas d'IP dans GLPI, on essaie par hostname
    if ip == "N/A" and computer.get("name"):
        try:
            ip = socket.gethostbyname(computer["name"])
        except Exception:
            ip = "N/A"

    statut_reseau = ping_host(ip)
    hostname      = resolve_hostname(ip)

    print(statut_reseau)

    return {
        "id":             computer.get("id", ""),
        "nom":            computer.get("name", "N/A"),
        "ip":             ip,
        "hostname":       hostname,
        "statut_reseau":  statut_reseau,
        "os":             computer.get("operatingsystems_id", "N/A"),
        "fabricant":      computer.get("manufacturers_id", "N/A"),
        "modele":         computer.get("computermodels_id", "N/A"),
        "serie":          computer.get("serial", "N/A"),
        "entite":         computer.get("entities_id", "N/A"),
        "localisation":   computer.get("locations_id", "N/A"),
        "utilisateur":    computer.get("users_id", "N/A"),
        "date_achat":     computer.get("buy_date", "N/A"),
        "date_modif":     computer.get("date_mod", "N/A")[:16].replace("T"," ")
                          if computer.get("date_mod") else "N/A",
        "commentaire":    computer.get("comment", "")
    }

# ── Export Excel ──────────────────────────────────────────
def export_inventaire_excel(equipements):
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Inventaire"

    # Styles
    bleu_fonce = PatternFill("solid", fgColor="1F4E79")
    bleu_moy   = PatternFill("solid", fgColor="2E75B6")
    vert       = PatternFill("solid", fgColor="C6EFCE")
    rouge      = PatternFill("solid", fgColor="FFC7CE")
    gris       = PatternFill("solid", fgColor="F2F2F2")
    blanc      = PatternFill("solid", fgColor="FFFFFF")

    bordure = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    # Titre
    ws.merge_cells("A1:N1")
    ws["A1"] = f"  Inventaire Réseau CIMAT Béni Mellal — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = bleu_fonce
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # En-têtes
    entetes = [
        "ID", "Nom machine", "Adresse IP", "Hostname réseau",
        "Statut réseau", "Système OS", "Fabricant", "Modèle",
        "N° Série", "Entité", "Localisation", "Utilisateur",
        "Date achat", "Dernière modif."
    ]
    largeurs = [6, 22, 16, 22, 14, 18, 16, 18, 16, 14, 16, 16, 14, 18]

    for col, (titre, larg) in enumerate(zip(entetes, largeurs), start=1):
        cell = ws.cell(row=2, column=col, value=titre)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = bleu_moy
        cell.alignment = Alignment(horizontal="center")
        cell.border    = bordure
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col)
        ].width = larg

    ws.row_dimensions[2].height = 20

    # Données
    for ligne, eq in enumerate(equipements, start=3):
        fill = gris if ligne % 2 == 0 else blanc
        valeurs = [
            eq["id"], eq["nom"], eq["ip"], eq["hostname"],
            eq["statut_reseau"], eq["os"], eq["fabricant"],
            eq["modele"], eq["serie"], eq["entite"],
            eq["localisation"], eq["utilisateur"],
            eq["date_achat"], eq["date_modif"]
        ]
        for col, val in enumerate(valeurs, start=1):
            cell = ws.cell(row=ligne, column=col, value=str(val))
            cell.border    = bordure
            cell.alignment = Alignment(horizontal="left", vertical="center")

            # Couleur statut réseau
            if col == 5:
                if "En ligne" in str(val):
                    cell.fill = vert
                elif "Hors ligne" in str(val):
                    cell.fill = rouge
                else:
                    cell.fill = fill
            else:
                cell.fill = fill

    # Feuille 2 — Résumé
    ws2 = wb.create_sheet("Résumé réseau")
    ws2.merge_cells("A1:C1")
    ws2["A1"] = " Résumé inventaire réseau"
    ws2["A1"].font      = Font(bold=True, size=12, color="FFFFFF")
    ws2["A1"].fill      = bleu_fonce
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 26

    en_ligne    = sum(1 for e in equipements if "En ligne"   in e["statut_reseau"])
    hors_ligne  = sum(1 for e in equipements if "Hors ligne" in e["statut_reseau"])
    inconnu     = sum(1 for e in equipements if "Inconnu"    in e["statut_reseau"])

    donnees_resume = [
        ("Total équipements",  len(equipements)),
        ("En ligne ",         en_ligne),
        ("Hors ligne ",       hors_ligne),
        ("Statut inconnu ❓",   inconnu),
    ]

    for i, (label, val) in enumerate(donnees_resume, start=3):
        ws2.cell(row=i, column=1, value=label).border = bordure
        c = ws2.cell(row=i, column=2, value=val)
        c.border = bordure
        c.font   = Font(bold=True)

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 10

    # Sauvegarde
    date_str    = datetime.now().strftime("%Y-%m-%d_%H-%M")
    nom_fichier = f"GLPI_Inventaire_{date_str}.xlsx"
    chemin      = os.path.join(os.path.dirname(__file__), nom_fichier)
    wb.save(chemin)
    print(f"\n Inventaire Excel créé : {chemin}")
    return chemin

# ── MAIN ──────────────────────────────────────────────────
if __name__ == "__main__":
    session = connect_glpi()
    if session:
        # Récupération ordinateurs + équipements réseau
        computers      = get_computers(session)
        net_devices    = get_network_devices(session)
        tous_equipements = computers + net_devices

        if tous_equipements:
            print(f"\n Enrichissement réseau de {len(tous_equipements)} équipement(s)…\n")
            enrichis = []
            for eq in tous_equipements:
                enrichis.append(enrichir_equipement(eq, session))
            export_inventaire_excel(enrichis)
        else:
            print("  Aucun équipement trouvé dans GLPI.")
            print(" Astuce : Ajoute des ordinateurs dans GLPI d'abord !")

        disconnect_glpi(session)