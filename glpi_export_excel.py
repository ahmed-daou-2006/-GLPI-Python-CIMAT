# glpi_export_excel.py — Export hebdomadaire des statistiques
# Projet CIMAT Béni Mellal

import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
import os

# ── Config ────────────────────────────────────────────────
GLPI_URL   = "http://192.168.112.129/glpi"
APP_TOKEN  = "Qzqd9JxK0k9vcF9daXxT4zIJKZGht2r20pHgaYR4"
USER_TOKEN = "gMJqxaSdpy79c9HQout8reOo1PBY0UfM15FkQGYs"

STATUTS = {
    1: "Nouveau",
    2: "En cours (Assigné)",
    3: "En cours (Planifié)",
    4: "En attente",
    5: "Résolu",
    6: "Clos"
}

PRIORITES = {
    1: "Très basse",
    2: "Basse",
    3: "Moyenne",
    4: "Haute",
    5: "Très haute",
    6: "Majeure"
}

TYPES = {
    1: "Incident",
    2: "Demande"
}

# ── Connexion ─────────────────────────────────────────────
def connect_glpi():
    headers = {
        "Content-Type":  "application/json",
        "Authorization": f"user_token {USER_TOKEN}",
        "App-Token":     APP_TOKEN
    }
    r = requests.get(f"{GLPI_URL}/apirest.php/initSession", headers=headers)
    data = r.json()
    if isinstance(data, list) or "session_token" not in data:
        print(f" Erreur connexion : {data}")
        return None
    print(f" Connecté — Session : {data['session_token'][:12]}…")
    return data["session_token"]

def disconnect_glpi(session_token):
    headers = {
        "Content-Type":  "application/json",
        "Session-Token": session_token,
        "App-Token":     APP_TOKEN
    }
    requests.get(f"{GLPI_URL}/apirest.php/killSession", headers=headers)
    print(" Session fermée.")

# ── Récupération des tickets ───────────────────────────────
def get_tickets(session_token):
    headers = {
        "Content-Type":  "application/json",
        "Session-Token": session_token,
        "App-Token":     APP_TOKEN
    }
    params = {
        "expand_dropdowns": True,
        "range": "0-999",
        "sort": "id",
        "order": "DESC"
    }
    r = requests.get(
        f"{GLPI_URL}/apirest.php/Ticket",
        headers=headers,
        params=params
    )
    if r.status_code != 200:
        print(f"❌ Erreur récupération tickets : {r.text}")
        return []
    tickets = r.json()
    print(f" {len(tickets)} ticket(s) récupéré(s).")
    return tickets

# ── Création du fichier Excel ──────────────────────────────
def export_excel(tickets):
    wb = openpyxl.Workbook()

    # ════════════════════════════════════
    # FEUILLE 1 — Liste des tickets
    # ════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Tickets"

    # Couleurs
    bleu_fonce  = PatternFill("solid", fgColor="1F4E79")
    bleu_clair  = PatternFill("solid", fgColor="D6E4F0")
    vert        = PatternFill("solid", fgColor="C6EFCE")
    rouge       = PatternFill("solid", fgColor="FFC7CE")
    orange      = PatternFill("solid", fgColor="FFEB9C")
    gris        = PatternFill("solid", fgColor="F2F2F2")

    bordure = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Titre principal
    ws1.merge_cells("A1:H1")
    ws1["A1"] = f" Rapport GLPI — CIMAT Béni Mellal — {datetime.now().strftime('%d/%m/%Y')}"
    ws1["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws1["A1"].fill      = bleu_fonce
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    # En-têtes colonnes
    entetes = ["ID", "Titre", "Type", "Statut", "Priorité",
               "Date ouverture", "Dernière modif.", "Entité"]
    for col, titre in enumerate(entetes, start=1):
        cell = ws1.cell(row=2, column=col, value=titre)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="2E75B6")
        cell.alignment = Alignment(horizontal="center")
        cell.border    = bordure

    # Largeurs colonnes
    largeurs = [8, 45, 12, 22, 14, 20, 20, 18]
    for col, larg in enumerate(largeurs, start=1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = larg

    # Données tickets
    for ligne, t in enumerate(tickets, start=3):
        fill_ligne = gris if ligne % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

        statut_id   = t.get("status", 0)
        priorite_id = t.get("priority", 0)

        valeurs = [
            t.get("id", ""),
            t.get("name", ""),
            TYPES.get(t.get("type", 1), "Inconnu"),
            STATUTS.get(statut_id, str(statut_id)),
            PRIORITES.get(priorite_id, str(priorite_id)),
            t.get("date", "")[:16].replace("T", " ") if t.get("date") else "",
            t.get("date_mod", "")[:16].replace("T", " ") if t.get("date_mod") else "",
            t.get("entities_id", "Entité racine")
                if isinstance(t.get("entities_id"), str)
                else "Entité racine"
        ]

        for col, val in enumerate(valeurs, start=1):
            cell = ws1.cell(row=ligne, column=col, value=val)
            cell.border    = bordure
            cell.alignment = Alignment(horizontal="left", vertical="center")

            # Couleur selon statut
            if col == 4:
                if statut_id == 1:   cell.fill = vert
                elif statut_id == 5: cell.fill = PatternFill("solid", fgColor="C6EFCE")
                elif statut_id == 6: cell.fill = gris
                else:                cell.fill = orange

            # Couleur selon priorité
            elif col == 5:
                if priorite_id >= 5:   cell.fill = rouge
                elif priorite_id == 4: cell.fill = orange
                elif priorite_id <= 2: cell.fill = bleu_clair
                else:                  cell.fill = fill_ligne
            else:
                cell.fill = fill_ligne

    # ════════════════════════════════════
    # FEUILLE 2 — Statistiques résumées
    # ════════════════════════════════════
    ws2 = wb.create_sheet("Statistiques")

    ws2.merge_cells("A1:C1")
    ws2["A1"] = " Statistiques hebdomadaires — CIMAT"
    ws2["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
    ws2["A1"].fill      = bleu_fonce
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 28

    # Comptage par statut
    ws2["A3"] = "Par statut"
    ws2["A3"].font = Font(bold=True, size=11)

    ws2["A4"] = "Statut"
    ws2["B4"] = "Nombre"
    for cell in [ws2["A4"], ws2["B4"]]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E75B6")
        cell.border = bordure

    comptage_statut = {}
    for t in tickets:
        s = STATUTS.get(t.get("status", 0), "Inconnu")
        comptage_statut[s] = comptage_statut.get(s, 0) + 1

    for i, (statut, count) in enumerate(comptage_statut.items(), start=5):
        ws2.cell(row=i, column=1, value=statut).border = bordure
        ws2.cell(row=i, column=2, value=count).border  = bordure

    # Comptage par priorité
    ws2["A12"] = "Par priorité"
    ws2["A12"].font = Font(bold=True, size=11)

    ws2["A13"] = "Priorité"
    ws2["B13"] = "Nombre"
    for cell in [ws2["A13"], ws2["B13"]]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E75B6")
        cell.border = bordure

    comptage_prio = {}
    for t in tickets:
        p = PRIORITES.get(t.get("priority", 0), "Inconnu")
        comptage_prio[p] = comptage_prio.get(p, 0) + 1

    for i, (prio, count) in enumerate(comptage_prio.items(), start=14):
        ws2.cell(row=i, column=1, value=prio).border = bordure
        ws2.cell(row=i, column=2, value=count).border = bordure

    # Total général
    ws2["A22"] = "TOTAL TICKETS"
    ws2["B22"] = len(tickets)
    ws2["A22"].font = Font(bold=True)
    ws2["B22"].font = Font(bold=True)

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 12

    # ── Sauvegarde ─────────────────────────────────────────
    date_str  = datetime.now().strftime("%Y-%m-%d_%H-%M")
    nom_fichier = f"GLPI_Rapport_{date_str}.xlsx"
    chemin      = os.path.join(os.path.dirname(__file__), nom_fichier)
    wb.save(chemin)
    print(f"\n Export Excel créé : {chemin}")
    return chemin

# ── MAIN ───────────────────────────────────────────────────
if __name__ == "__main__":
    session = connect_glpi()
    if session:
        tickets = get_tickets(session)
        if tickets:
            export_excel(tickets)
        disconnect_glpi(session)